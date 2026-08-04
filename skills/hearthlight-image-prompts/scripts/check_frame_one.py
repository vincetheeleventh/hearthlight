#!/usr/bin/env python3
"""FRAME-ONE LAW validator. Scans a shot list's Still column for temporal language.

A still is ONE INSTANT. An image generator has no concept of "after" — it renders
every clause simultaneously. A Still spec containing a sequence produces a frame in
which both states happen at once (the yugioh shot-2 failure: the wife handing over
dog tags while the father is still tying his boot).

Usage:  python3 check_frame_one.py <shotlist.xlsx> [--sheet "Shot List"]
Exit 1 if any violation is found. Run BEFORE spending a generation.
"""
import sys, re, argparse

BANNED = re.compile(
    r'\b(after|then|once|next|meanwhile|begins?\s+to|starts?\s+to|about\s+to|'
    r'as\s+(?:he|she|they|it|the|his|her)|enters?\s+(?:frame|the)|comes?\s+down|'
    r'lowers?\s+.{0,30}?\s+into|reaches?\s+(?:for|out)|turns?\s+and|straightens?|'
    r'rises?|stands?\s+up|sits?\s+down|and\s+takes?\s+(?:them|it))\b', re.I)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("path"); ap.add_argument("--sheet", default="Shot List")
    ap.add_argument("--col", default="Still (frame one)")
    a = ap.parse_args()
    import openpyxl
    ws = openpyxl.load_workbook(a.path)[a.sheet]
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    col = hdr.get(a.col) or hdr.get("Visual Description")
    if not col:
        print(f"FAIL: no '{a.col}' or 'Visual Description' column"); return 2
    POSSESSIVE = re.compile(
        r"\b\w+(?:s|e)'s\b(?=\s+(?:hand|face|shoulder|arm|forearm|sleeve|head|eye|body|open))"
        r"|\b\w+s's\b")
    bad, gram = [], []
    for r in range(2, ws.max_row + 1):
        shot = ws.cell(r, 1).value
        if shot is None: continue
        txt = str(ws.cell(r, col).value or "")
        hits = sorted({m.group(0).lower() for m in BANNED.finditer(txt)})
        if hits:
            bad.append((shot, ws.cell(r, 2).value, hits))
        g = sorted({m.group(0) for m in POSSESSIVE.finditer(txt)})
        if g:
            gram.append((shot, ws.cell(r, 2).value, g))
    if gram:
        print("POSSESSIVE BREAK — a signature string was used in the possessive:\n")
        for shot, title, g in gram:
            print(f"  shot {shot} ({title}): {g}")
        print("\nSignature strings END IN NOUNS ('tired eyes', 'tan boots', 'a band-aid on one knee'),")
        print("so X's is always ungrammatical. Write 'the hand of {X}', never '{X}'s hand'.\n")
    if not bad and not gram:
        print(f"FRAME-ONE LAW: PASS — {ws.max_row-1} stills, no temporal language."); return 0
    print("FRAME-ONE LAW: FAIL — these stills describe change over time:\n")
    for shot, title, hits in bad:
        print(f"  shot {shot} ({title}): {hits}")
    print("\nEach is a two-state shot. Resolve by: splitting into two stills (preferred when NEW")
    print("information enters frame), conditioning on frame one only (when the change is movement")
    print("of something already visible), or conditioning on the later state (loses the reveal).")
    return 1

if __name__ == "__main__":
    sys.exit(main())
